from flask import Flask, jsonify, request, send_from_directory, render_template, render_template_string, session, redirect, url_for
from functools import wraps
import json, os, io, secrets
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=BASE_DIR)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64 MB
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
SITE_PASSWORD = os.environ.get('SITE_PASSWORD', 'nikou')
DATA_DIR = '/tmp' if os.environ.get('RENDER') else BASE_DIR
DATA_FILE = os.path.join(DATA_DIR, 'data.json')
PREP_STATE_FILE = os.path.join(DATA_DIR, 'prep_state.json')
SKIP_SHEETS = {'EXEMPLE', 'Flipper', 'last stop'}

LOGIN_HTML = '''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Login</title>
<style>
  body{margin:0;display:flex;align-items:center;justify-content:center;min-height:100vh;background:#0f1923;font-family:sans-serif}
  .box{background:#1a2634;padding:2.5rem 3rem;border-radius:12px;box-shadow:0 4px 24px #0008;text-align:center;min-width:280px}
  h2{color:#fff;margin:0 0 1.5rem;font-size:1.4rem;letter-spacing:.05em}
  input{width:100%;box-sizing:border-box;padding:.7rem 1rem;border:1px solid #304057;border-radius:8px;background:#0f1923;color:#fff;font-size:1rem;margin-bottom:1rem;outline:none}
  input:focus{border-color:#4a90e2}
  button{width:100%;padding:.75rem;background:#4a90e2;color:#fff;border:none;border-radius:8px;font-size:1rem;cursor:pointer;font-weight:600}
  button:hover{background:#357abd}
  .err{color:#e25555;margin-bottom:.75rem;font-size:.9rem}
</style></head>
<body><div class="box">
  <h2>nikou</h2>
  {% if error %}<div class="err">Wrong password</div>{% endif %}
  <form method="post">
    <input type="password" name="password" placeholder="Password" autofocus>
    <button type="submit">Enter</button>
  </form>
</div></body></html>'''


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = False
    if request.method == 'POST':
        if request.form.get('password') == SITE_PASSWORD:
            session['logged_in'] = True
            next_url = request.args.get('next') or '/'
            return redirect(next_url)
        error = True
    return render_template_string(LOGIN_HTML, error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


MAP_MODES = {
    'shooting star': 'Bounty', 'snake prairie': 'Bounty', 'layer cake': 'Bounty',
    'dry season': 'Bounty', 'dry sason': 'Bounty',
    'super beach': 'Brawl Ball', 'pinhole punt': 'Brawl Ball', 'center stage': 'Brawl Ball',
    'triple dribble': 'Brawl Ball', 'sunny soccer': 'Brawl Ball', 'sneaky fields': 'Brawl Ball',
    'pinball dreams': 'Brawl Ball',
    'hard rock mine': 'Gem Grab', 'double swoosh': 'Gem Grab', 'deathcap trap': 'Gem Grab',
    'undermine': 'Gem Grab', 'gem fort': 'Gem Grab',
    'safe zone': 'Heist', 'kaboom canyon': 'Heist', 'bridge too far': 'Heist',
    'hot potato': 'Heist', 'pit stop': 'Heist',
    'open business': 'Hot Zone', 'dueling beetles': 'Hot Zone', 'ring of fire': 'Hot Zone',
    'goldarm gulch': 'Knockout', 'belles rock': 'Knockout', 'out in the open': 'Knockout',
    'hideout': 'Knockout',
}


def get_map_mode(map_name):
    key = map_name.lower().strip()
    if key in MAP_MODES:
        return MAP_MODES[key]
    for k, v in MAP_MODES.items():
        if key.startswith(k[:8]) or k.startswith(key[:8]):
            return v
    return 'Other'


def compute_stats():
    if not os.path.exists(DATA_FILE):
        return [], {}, {}, {}, {}
    with open(DATA_FILE) as f:
        data = json.load(f)

    brawler_record = defaultdict(lambda: [0, 0])   # name -> [wins, games]
    synergy = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    matchup = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    maps_by_mode = defaultdict(list)

    for entry in data:
        mode = get_map_mode(entry['map'])
        if entry['map'] not in maps_by_mode[mode]:
            maps_by_mode[mode].append(entry['map'])

        for m in entry['matches']:
            blue = [m[k] for k in ('pick2', 'pick3', 'pick6') if m.get(k)]
            red  = [m[k] for k in ('pick1', 'pick4', 'pick5') if m.get(k)]
            try:
                bs = float(m.get('score_blue') or 0)
                rs = float(m.get('score_red') or 0)
            except (ValueError, TypeError):
                continue
            blue_won = bs > rs
            red_won  = rs > bs

            for brawler in blue:
                brawler_record[brawler][1] += 1
                if blue_won: brawler_record[brawler][0] += 1
            for brawler in red:
                brawler_record[brawler][1] += 1
                if red_won: brawler_record[brawler][0] += 1

            # same-team synergy
            for team, won in ((blue, blue_won), (red, red_won)):
                for i, b1 in enumerate(team):
                    for b2 in team[i+1:]:
                        synergy[b1][b2][1] += 1
                        synergy[b2][b1][1] += 1
                        if won:
                            synergy[b1][b2][0] += 1
                            synergy[b2][b1][0] += 1

            # cross-team matchups
            for b1 in blue:
                for b2 in red:
                    matchup[b1][b2][1] += 1
                    matchup[b2][b1][1] += 1
                    if blue_won:
                        matchup[b1][b2][0] += 1
                    elif red_won:
                        matchup[b2][b1][0] += 1

    def pct(w, t): return round(w / t * 100, 1) if t else 50.0

    all_brawlers = sorted(brawler_record.keys())
    meta_scores  = {b: pct(w, t) for b, (w, t) in brawler_record.items()}
    synergy_data = {b1: {b2: pct(v[0], v[1]) for b2, v in b2s.items()} for b1, b2s in synergy.items()}
    matchup_data = {b1: {b2: pct(v[0], v[1]) for b2, v in b2s.items()} for b1, b2s in matchup.items()}

    return all_brawlers, dict(maps_by_mode), meta_scores, synergy_data, matchup_data


# ── Excel parsing (for upload) ──────────────────────────────────────────────

def clean(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan', 'none') else s


def find_header_row(rows):
    for i, row in enumerate(rows[:10]):
        vals = [clean(c).lower() for c in row]
        if 'blue' in vals or 'team' in vals:
            return i
    return -1


def parse_workbook(wb):
    result = []
    print(f'[parse] sheets: {wb.sheetnames}')
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        if not rows:
            continue
        hdr_idx = find_header_row(rows)
        if hdr_idx == -1:
            print(f'  {sheet_name}: header not found, first row = {rows[0][:5]}')
            continue
        print(f'  {sheet_name}: header at row {hdr_idx}')
        matches = []
        for row in rows[hdr_idx + 1:]:
            while len(row) < 17:
                row.append(None)
            def v(i, _r=row):
                return clean(_r[i]) if i < len(_r) else ''
            blue, red = v(1), v(2)
            if not blue and not red:
                continue
            if not any(v(i) for i in range(9, 15)):
                continue
            sb, sr = v(15), v(16)
            if '/' in sb:
                parts = sb.split('/')
                sb, sr = parts[0].strip(), parts[1].strip()
            raw_date = row[0]
            date_str = ''
            if raw_date is not None and raw_date != '':
                if isinstance(raw_date, datetime):
                    date_str = raw_date.strftime('%d/%m/%Y')
                elif str(raw_date).strip():
                    date_str = str(raw_date).strip()
            matches.append({
                'blue': blue, 'red': red, 'date': date_str,
                'ban2_1': v(3), 'ban2_2': v(4), 'ban2_3': v(5),
                'ban1_1': v(6), 'ban1_2': v(7), 'ban1_3': v(8),
                'pick1': v(9), 'pick2': v(10), 'pick3': v(11),
                'pick4': v(12), 'pick5': v(13), 'pick6': v(14),
                'score_blue': sb, 'score_red': sr,
            })
        print(f'    -> {len(matches)} matches')
        if matches:
            result.append({'map': sheet_name.strip(), 'matches': matches})
    print(f'[parse] total maps: {len(result)}')
    return result


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route('/static/images/<path:path>')
@login_required
def serve_images(path):
    response = send_from_directory(BASE_DIR, path)
    response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    return response


@app.route('/api/data')
@login_required
def get_data():
    if not os.path.exists(DATA_FILE):
        return jsonify([])
    with open(DATA_FILE) as f:
        return jsonify(json.load(f))


@app.errorhandler(413)
def too_large(e):
    return jsonify({'ok': False, 'error': 'File too large (max 64MB)'}), 413

@app.errorhandler(500)
def server_error(e):
    return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])
@login_required
def upload():
    print('[upload] request received')
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file provided'}), 400
    try:
        file_bytes = io.BytesIO(request.files['file'].read())
        wb = load_workbook(file_bytes, read_only=True, data_only=True)
        data = parse_workbook(wb)
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f)
        return jsonify({'ok': True, 'maps': len(data)})
    except Exception as e:
        print(f'[upload error] {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/prep-state', methods=['GET'])
@login_required
def get_prep_state():
    if not os.path.exists(PREP_STATE_FILE):
        return jsonify({})
    with open(PREP_STATE_FILE) as f:
        return jsonify(json.load(f))


@app.route('/api/prep-state', methods=['POST'])
@login_required
def save_prep_state():
    data = request.get_json()
    with open(PREP_STATE_FILE, 'w') as f:
        json.dump(data, f)
    return jsonify({'ok': True})


@app.route('/prep')
@login_required
def prep():
    all_brawlers, maps_by_mode, meta_scores, synergy_data, matchup_data = compute_stats()
    return render_template('prep.html',
        all_brawlers=all_brawlers,
        maps_by_mode=maps_by_mode,
        meta_scores=meta_scores,
        synergy_data=synergy_data,
        matchup_data=matchup_data,
    )


IMAGE_DIRS = ('brawlers/', 'maps/', 'modes/')

@app.route('/', defaults={'path': 'draft.html'})
@app.route('/<path:path>')
@login_required
def serve_static(path):
    response = send_from_directory(BASE_DIR, path)
    if path.startswith(IMAGE_DIRS):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    return response


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    app.run(port=port)
